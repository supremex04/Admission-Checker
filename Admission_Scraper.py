'''
************************************************************
Steps:
1) Convert the priority list pdf file into an excel file here: https://www.adobe.com/acrobat/online/pdf-to-excel.html
2) Check the hashed ('#') out texts for further instructions
************************************************************
'''
#type "pip install openpyxl" in cmd before running this program
import openpyxl
import os
#below provide the directory path where the converted excel file lies
os.chdir('c:\\users\\Ram\\Downloads')
#below provide the name of the converted excel file
wb = openpyxl.load_workbook('fromadobe.xlsx')
#open the excel file, check the lower left corner, if you see 'Table 1', leave the below line as it it.
#if you see 'Sheet1' or any other name for that sheet on the lower left corner, type that name below
sheet = wb.get_sheet_by_name('Table 1')
newb = openpyxl.Workbook()
s = newb.get_sheet_by_name('Sheet')
z = 1
P1 = int(input("Your first priority subject code: "))
P2 = int(input("Your second priority subject code: "))
rank = int(input("Your rank: "))
s.cell(row=1, column=1).value = 'Name'
s.cell(row=1, column=2).value = 'Rank'
s.cell(row=1, column=3).value = 'P1'
s.cell(row=1, column=4).value = 'P2'
s.cell(row=1, column=5).value = 'P3'
s.cell(row=1, column=6).value = 'P4'
s.cell(row=1, column=7).value = 'P5'

for i in range(1, 1973):
    # checking for int as there seems to be random open cells in the excel file
    if type(sheet.cell(row=i, column=7).value) == int:
        if sheet.cell(row=i, column=7).value < rank:
            if (sheet.cell(row=i, column=8).value == (P1 or P2)) or (sheet.cell(row=i, column=9).value == (P1 or P2)):
                z = z + 1
                print(i, 'done', z)
                s.cell(row=(z),
                       column=2).value = int(sheet.cell(row=i, column=7).value)
                s.cell(row=(z),
                       column=1).value = sheet.cell(row=i, column=3).value
                s.cell(row=(z),
                       column=3).value = sheet.cell(row=i, column=8).value
                s.cell(row=(z),
                       column=4).value = sheet.cell(row=i, column=9).value
                s.cell(row=(z),
                       column=5).value = sheet.cell(row=i, column=10).value
                s.cell(row=(z),
                       column=6).value = sheet.cell(row=i, column=11).value
                s.cell(row=(z),
                       column=7).value = sheet.cell(row=i, column=12).value


# here input the folder you want to save the new excel file in
os.chdir('c:\\Python_Learn')
#name of the new excel file will be YourList.xlsx
newb.save('YourList.xlsx')
#You're all good to go!
